from apiclient import discovery
from oauth2client import client
from oauth2client.service_account import ServiceAccountCredentials
import webbrowser
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient import errors as googleapierrors
import io
import csv
import os
import openpyxl
import json


class GoogleAPI:

    def __init__(self, credentials_file=None, credentials_json=None, enable_drive=True, enable_sheets=True, auto_ownership=False, default_owner_email=''):
        self.CREDENTIALS = credentials_file
        self.CREDENTIALS_JSON = credentials_json
        self.auth_type = None
        self.service_email = None

        # if the authentication type is a service account, allow the user to specify if
        #   we should automatically make them the owner of every file we upload
        self.auto_owner = auto_ownership
        self.default_owner_email = default_owner_email

        self.drive_service = None
        self.sheet_service = None
        self._sheet_version = 'v4'
        self._drive_version = 'v3'

        # do we want to enable Google Sheets API?
        self.sheets_enabled = enable_sheets
        self._test_sheets_on_auth = False

        # do we want to enable Google Drive API?
        self.drive_enabled = enable_drive
        self._test_drive_on_auth = False

        # make sure drive is enabled if sheets is enabled
        if self.sheets_enabled and not self.drive_enabled:
            raise AttributeError("Many functions for Sheets require support from the Drive API.  Please enable the "
                                 "Drive feature if you want to use Sheets.")

    def authorize(self, authentication_type=None, browser=None):
        """
        Makes an attempt to get authorization to the enabled Google APIs
        :param authentication_type: Can be 'AUTH_TYPE_BROWSER' or the string 'web', OR, 'AUTH_TYPE_SERVICE_ACCOUNT' or the string 'service_account'
        :param browser: Can be 'BROWSER_GOOGLE_CHROME' or 'BROWSER_FIREFOX', or the strings 'chrome'|'firefox'.
                        If not passed in, the system default will be used.
        :return: True if all enabled drives are authenticated.  False if an error occurs.
        """
        if not self.CREDENTIALS and not self.CREDENTIALS_JSON:
            raise FileNotFoundError('No credentials file has been set.  See: \n\t'
                                    'https://cloud.google.com/genomics/downloading-credentials-for-api-access\n'
                                    'for instructions on how to obtain this file. ')
        credentials = None
        if authentication_type == 'web':
            flow = client.flow_from_clientsecrets(
                self.CREDENTIALS,
                scope='https://www.googleapis.com/auth/drive',
                redirect_uri='urn:ietf:wg:oauth:2.0:oob')

            auth_uri = flow.step1_get_authorize_url()
            if browser:
                webbrowser.get(browser).open(auth_uri)
            else:
                webbrowser.open(auth_uri)

            auth_code = input('Enter the auth code: ')
            credentials = flow.step2_exchange(auth_code)

        if authentication_type == 'service_account':
            # A complete list of scopes can be found at: https://developers.google.com/identity/protocols/googlescopes#drivev3
            SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/drive.appdata',
                      'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/drive.metadata',
                      'https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive.file']
            if self.CREDENTIALS:
                credentials = ServiceAccountCredentials.from_json_keyfile_name(
                    self.CREDENTIALS, scopes=SCOPES)
                creds = open(self.CREDENTIALS, 'r')
                cred_obj = json.load(creds)
                creds.close()
                self.service_email = cred_obj['client_email']
            if self.CREDENTIALS_JSON:
                credentials = ServiceAccountCredentials.from_json_keyfile_dict(
                    self.CREDENTIALS_JSON, scopes=SCOPES)
                self.service_email = self.CREDENTIALS_JSON['client_email']

        if authentication_type != 'web' and authentication_type != 'service_account':
            raise AttributeError('Unsupported authentication type: `' + authentication_type + '`')

        self.auth_type = authentication_type

        # is drive enabled?
        if self.drive_enabled:
            self.drive_service = discovery.build('drive', self._drive_version, credentials=credentials)

            # test the API
            if self._test_drive_on_auth:
                try:
                    results = self.drive_service.files().list(
                        pageSize=10, fields="nextPageToken, files(id, name)").execute()

                except googleapierrors.HttpError as e:
                    if e.resp.status == 403:
                        raise ConnectionError(
                            "The Google Drive API rejected your request.  It is likely the API is not enabled, or you have exceeded your limit.\n\tMake sure the Drive AIP is enabled at: https://console.cloud.google.com/apis/library/drive.googleapis.com/?q=Drive")
                    else:
                        raise e
        # is sheets enabled?
        if self.sheets_enabled:
            self.sheet_service = discovery.build('sheets', version=self._sheet_version, credentials=credentials)

            # test the API
            if self._test_sheets_on_auth:
                try:
                    # try creating a new sheet
                    request = self.sheet_service.spreadsheets().create(body={})
                    response = request.execute()

                    # and removing it
                    self.drive_service.files().delete(fileId=response['spreadsheetId']).execute()

                except googleapierrors.HttpError as e:
                    if e.resp.status == 403:
                        raise ConnectionError(
                            "The Google Drive API rejected your request.  It is likely the API is not enabled, or you have exceeded your limit.\n\tMake sure the Sheets AIP is enabled at: https://console.cloud.google.com/apis/library/sheets.googleapis.com/?q=Sheet")
                    else:
                        raise e
        return True


    # def create_spreadsheet(self):
    #     ##TODO: Fill out
    #     pass

    def revoke_permissions(self, file_id, user_email):
        """
        Removes all permissions/sharing for the given file from the email address specified.
        :param file_id: File ID of the file to revoke permissions for.
        :param user_email: Email address of the user to revoke.
        :return: List of permissions it revoked (email/permission ID)
        """
        # first get a list of permissions for this file
        permissions = self.get_file_permissions(file_id=file_id)
        permissions_removed = []
        for permission in permissions:
            if permission['emailAddress'] == user_email:
                self.drive_service.permissions().delete(fileId=file_id, permissionId=permission['id']).execute()
                permissions_removed.append(permission)
        return permissions_removed

    def share_document(self, file_id, recipient_email, share_permissions='reader'):
        """
        Shares a file with the user specified email address to have access defined by the share_permissions attribute
        :param file_id: File ID you want to share
        :param recipient_email: Email address of the user you want to share with
        :param share_permissions: Any of the known PERMISSION_CONSTANTS of reader/writer/owner.
        :return: True if the share was successful
        """
        batch = self.drive_service.new_batch_http_request()
        user_permission = {}
        if share_permissions == 'reader' or share_permissions == 'writer':
            user_permission = {
                'type': 'user',
                'role': share_permissions,
                'emailAddress': recipient_email
            }
            create = self.drive_service.permissions().create(
                fileId=file_id,
                body=user_permission,
                fields='id'
            )
        else:
            user_permission = {
                'type': 'user',
                'role': 'owner',
                'transferOwnership': True,
                'emailAddress': recipient_email
            }
            create = self.drive_service.permissions().create(
                fileId=file_id,
                body=user_permission,
                transferOwnership=True,
                fields='id'
            )
        response = None
        try:
            response = create.execute()
        except googleapierrors.HttpError as e:
            print("=====")
            print(e.content)
            if e.resp.status == 404 and self.auth_type == 'service_account':
                raise FileNotFoundError("The service account you are authorizing as could not find the file.  "
                                        "It is likely "
                                        "you have not shared the file with the service account.\nPlease be sure to "
                                        "share "
                                        "the file with the service account email: " + self.service_email)
            if e.resp.status == 403 and self.auth_type == 'service_account':
                raise PermissionError("You cannot transfer ownership using this service account before first setting "
                                      "the service account to the owner of the document.  \nPlease manually set the "
                                      "owner of this file to the service account email: "+self.service_email)
            if e.resp.status == 400:
                raise PermissionError("Looks like you're trying to transfer ownership to another domain, and sadly, "
                                      "Google won't let you do that. Boo... :(")
        return response

    def get_file_permissions(self, file_id):
        """
        Returns a list of all permissions on the given File ID
        :param file_id: File ID you want to get the permissions for.
        :return: List of permissions as JSON object.
        """
        permissions = self.drive_service.permissions().list(fileId=file_id).execute()['permissions']
        permission_list = []
        for permission in permissions:
            permission_list.append(self.drive_service.permissions().get(fileId=file_id, permissionId=permission['id'],
                                                                        fields='emailAddress, id').execute())

        return permission_list

    def replace_spreadsheet_with_csv(self, spreadsheet_id, csv_file_location, input_type='USER_ENTERED', tab_index=None,
                                     tab_name=None):
        """
        First clears the specified sheet (tab-supported by name or index), then uploads the CSV contents to the default
        "A1" range.
        :param spreadsheet_id: Google Sheets File ID
        :param csv_file_location: Location of the CSV file to upload and replace the existing sheet with
        :param input_type: Any of the known INPUT_TYPE_CONSTANTS; determines if values should be interpreted or raw
        :param tab_index: Optional tab index if the target is not the first tab/sheet
        :param tab_name: Optioanl tab name if the target is not the first tab/sheet
        :return:
        """
        # first clear the spreadsheet
        self.clear_spreadsheet(spreadsheet_id=spreadsheet_id, tab_index=tab_index, tab_name=tab_name)

        # now upload this sheet
        return self.append_file_to_spreadsheet(spreadsheet_id=spreadsheet_id, csv_file=csv_file_location, tab_name=tab_name,
                                               tab_index=tab_index, input_type=input_type)

    def replace_spreadsheet_with_rows(self, spreadsheet_id, row_data, input_type='USER_ENTERED', tab_index=None,
                                      tab_name=None):
        """
        First clears the specified sheet (tab-supported by name or index), then uploads the row data to the default
        "A1" range.
        :param spreadsheet_id: Google Sheets File ID
        :param row_data: List of rows (lists) containing data to upload
        :param input_type: Any of the known INPUT_TYPE_CONSTANTS; determines if values should be interpreted or raw
        :param tab_index: Optional tab index if the target is not the first tab/sheet
        :param tab_name: Optioanl tab name if the target is not the first tab/sheet
        :return:
        """
        # first clear the spreadsheet
        self.clear_spreadsheet(spreadsheet_id=spreadsheet_id, tab_index=tab_index, tab_name=tab_name)

        # now upload this sheet
        return self.append_rows_to_spreadsheet(spreadsheet_id=spreadsheet_id, row_data=row_data, tab_name=tab_name,
                                               tab_index=tab_index, input_type=input_type)


    def clear_spreadsheet(self, spreadsheet_id, tab_index=None, tab_name=None):
        """
        Clears a specific spreadsheet tab.  If no tab is specified, the first tab/sheet is cleared.
        NOTE: Cell formatting is preserved in this function.
        :param spreadsheet_id: Google Sheets File ID
        :param tab_index: which tab by index to clear.  Default to the first tab (0).
        :param tab_name: which tab by name to clear.  Default to the first tab.
        :return:
        """
        doc = self.sheet_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = [s['properties'] for s in doc['sheets']]
        sheet = sheets[0]
        if tab_index:
            sheet = sheets[tab_index]
        if tab_name:
            for tab in sheets:
                if tab['title'] == tab_name:
                    sheet = tab

        requests = []
        requests.append({
            "updateCells": {
                "range": {
                    "sheetId": sheet['sheetId']
                },
                'fields': 'userEnteredValue'
            }
        })

        body = {
            'requests': requests
        }
        response = self.sheet_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id,
                                                                 body=body).execute()
        return response

    def download_spreadsheet(self, spreadsheet_id, format, tab_index=None, tab_name=None):
        """
        Fetches the given Google Sheet in the given format as a io.ByteIO object.  USE THIS if you want to download
        the sheet into memory (no need for writing to disk).
        :param spreadsheet_id: Google Sheets File ID
        :param format: One of the known FORMAT_CONTSTANTS to download it as.  Use the `._get_format_extension()`
                        function to determine appropriate extension of the given format.
        :param tab_index: Only used if format is CSV or TSV - downloads the specific tab by index.
        :param tab_name: Only used if format is CSV or TSV - downloads the specific tab by name.
        :return: io.BytesIO object containing file data.
        """
        if not self.drive_service:
            raise ConnectionError("Drive service has not been turned on or not yet authenticated.  "
                                  "Please run the `authenticate()` function before attempting to manage files.")

        if format == FORMAT_CSV and tab_index is None and tab_name is None:
            print("Warning: CSV downloads without specifying a tab will result in a CSV of only the first tab.")

        if format == FORMAT_TSV and tab_index is None and tab_name is None:
            print("Warning: TSV downloads without specifying a tab will result in a TSV of only the first tab.")

        download_format = format
        if format == FORMAT_TSV or format == FORMAT_CSV:
            download_format = FORMAT_MS_EXCEL

        # Other formats supported:
        #   https://developers.google.com/drive/v3/web/manage-downloads#downloading_google_documents
        request = self.drive_service.files().export_media(fileId=spreadsheet_id,
                                                          mimeType=download_format)

        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            # print "Download %d%%." % int(status.progress() * 100)
        fh.seek(0)

        if (format == FORMAT_CSV or format == FORMAT_TSV) and (tab_index != None or tab_name):
            workbook = openpyxl.load_workbook(filename=fh)
            wb_sheet = None
            if tab_index != None:
                wb_sheet = workbook.worksheets[tab_index]
            if tab_name != None:
                wb_sheet = workbook[tab_name]
            
            # default to tab zero (first sheet)
            if tab_name is None and tab_index is None:
                wb_sheet = workbook.worksheets[0]

            csv_bytes = io.BytesIO()
            for row in wb_sheet.iter_rows():
                csv_row = []
                for cell in row:
                    csv_row.append(cell.value)
                csv_bytes.write(csv_row)
            return csv_bytes
        else:
            return fh

    def download_spreadsheet_to_file(self, spreadsheet_id, download_location, format, tab_index=None, tab_name=None):
        """
        Exports a given spreadsheet to the location specified.
        :param spreadsheet_id: File ID of the Google Sheet
        :param download_location: Where to save the file to.
        :param format: One of any known FORMAT_CONSTANTS.  File extension is derived from this.
        :param tab_index: Used only if exporting to CSV or TSV - defines which sheet/tab to download by sheet index.
                            (Hint: 0 = first sheet, 1 = second sheet, etc.)
        :param tab_name: Used only if exporting to CSV or TSV - defines which sheet/tab to download by sheet title.
        :return: File location of the downloaded file.
        """

        output_file = os.path.join(download_location, self.get_file_name(file_id=spreadsheet_id))
        # is the format TSV or CSV and a tab indicator given?
        if (format == FORMAT_CSV or format == FORMAT_TSV) and (tab_index != None or tab_name):
            fh = self.download_spreadsheet(spreadsheet_id=spreadsheet_id, format=FORMAT_MS_EXCEL, tab_index=tab_index,
                                           tab_name=tab_name)
            workbook = openpyxl.load_workbook(filename=fh)
            wb_sheet = None
            if tab_index != None:
                wb_sheet = workbook.worksheets[tab_index]
            else:
                wb_sheet = workbook[tab_name]
            
            csv_file = open(os.path.join(download_location, os.path.basename(output_file)+'.csv'), 'w')
            csv_writer = csv.writer(csv_file)
            for row in wb_sheet.iter_rows():
                csv_row = []
                for cell in row:
                    csv_row.append(cell.value)
                csv_writer.writerow(csv_row)
            csv_file.close()

            output_file = os.path.join(download_location, os.path.basename(output_file)+'.csv')

        else:
            # output the file instead
            fh = self.download_spreadsheet(spreadsheet_id=spreadsheet_id, format=format, tab_index=tab_index,
                                           tab_name=tab_name)
            output_file += self._get_format_extension(format)
            fout = open(output_file, 'wb')
            fout.write(fh.read())
            fout.close()

        return output_file

    def _get_format_extension(self, format):
        """
        Determines the appropriate file extension based on a known format
        :param format: One of the known FORMAT_CONSTANTS
        :return: string of extension proceeding a period
        """
        if format == FORMAT_CSV:
            return '.csv'
        if format == FORMAT_MS_EXCEL:
            return '.xlsx'
        if format == FORMAT_TSV:
            return '.tsv'
        if format == FORMAT_HTML:
            return '.zip'
        if format == FORMAT_OPEN_OFFICE_SHEET:
            return '.ods'
        if format == FORMAT_PDF:
            return '.pdf'
        return ''

    def get_file_name(self, file_id):
        return self.drive_service.files().get(fileId=file_id,
                                                 fields='name').execute()['name']

    def upload_spreadsheet(self, spreadsheet_location, parent_id=None):
        """
        Uploads a spreadsheet file to Google Sheets.  This simply adds the appropirate "mime type" parameter to tell Google to put it in Sheets.
        See: https://developers.google.com/drive/v3/web/mime-types
        :param spreadsheet_location: Location of the spreadsheet to upload
        :param parent_id: Optional parent folder ID
        :return: File object in JSON
        """
        return self.upload_file(spreadsheet_location, parent_id=parent_id,
                                mime_type='application/vnd.google-apps.spreadsheet')

    def upload_document(self, document_location, parent_id=None):
        """
        Uploads a document file to Google Sheets.  This simply adds the appropirate "mime type" parameter to tell Google to put it in Sheets.
        See: https://developers.google.com/drive/v3/web/mime-types
        :param document_location: Location of the document to upload
        :param parent_id: Optional parent folder ID
        :return: File object in JSON
        """
        return self.upload_file(document_location, parent_id=parent_id,
                                mime_type='application/vnd.google-apps.document')

    def delete_spreadsheet(self, spreadsheet_id=None):
        """
        Deletes a spreadsheet file by calling the "delete_file" function and passing the ID along.
        :param spreadsheet_id: File ID of the spreadsheet (specified as "spreadsheetID" in the spreadsheet object)
        :return: True if the file is deleted
        """
        if not spreadsheet_id:
            raise AttributeError("No spreadsheetID given.  Please specify the spreadsheetID/File ID.")
        return self.delete_file(file_id=spreadsheet_id)

    def create_folder(self, folder_name='Unnamed Folder', parent_folder_id=None):
        """
        Creates a folder of the name passed in.  If no parent folder ID is given, the folder will appear at the root of your Google Drive space.
        :param folder_name: Name of the folder you want to create.  If none is given, it will be called "Unnamed Folder"
        :param parent_folder_id: Optional parent folder ID.
        :return: Folder JSON Object meta data.
        """
        if not self.drive_service:
            raise ConnectionError("Drive service has not been turned on or not yet authenticated.  "
                                  "Please run the `authenticate()` function before attempting to manage files.")

        file_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        if parent_folder_id:
            file_metadata['parents'] = [parent_folder_id]

        file = self.drive_service.files().create(body=file_metadata).execute()

        # if this is a service account and auto-ownership is True
        if self.auth_type == 'service_account' and self.auto_owner:
            # move ownership to the default email
            if self.default_owner_email == '':
                raise AttributeError("No default owner email was set.  Cannot change ownership!")
            self.share_document(file_id=file['id'], recipient_email=self.default_owner_email, share_permissions='owner')
        return file

    def delete_file(self, file_id=None):
        """
        Deletes the file of `file_id` from your Google Drive space.
        :param file_id: ID of the file you want to delete
        :return: True if the deletion is successful
        """
        if not self.drive_service:
            raise ConnectionError("Drive service has not been turned on or not yet authenticated.  "
                                  "Please run the `authenticate()` function before attempting to manage files.")
        try:
            self.drive_service.files().delete(fileId=file_id).execute()
        except googleapierrors.HttpError as e:
            raise e
        return True

    def upload_file(self, file_location=None, parent_id=None, mime_type=None):
        """
        Uploads a file to your Google Drive space.  If a parent ID is given, it will upload it to that folder.
        :param file_location: Local path of the file to upload.
        :param parent_id: Folder ID to upload the file to.  If none is given, it will upload it to the root Google Drive space.
        :param mime_type: A known mimeType for Google to use to convert an uploaded file to Docs, Sheets, or Slides.  See list of supported mimeTypes: https://developers.google.com/drive/v3/web/mime-types
        :return:
        """
        if not file_location:
            raise AttributeError("No file location was given.")

        if not self.drive_service:
            raise ConnectionError("Drive service has not been turned on or not yet authenticated.  "
                                  "Please run the `authenticate()` function before attempting to manage files.")
        file_metadata = {
            'name': os.path.basename(file_location)
        }

        if parent_id:
            file_metadata['parents'] = [parent_id]

        if mime_type:
            file_metadata['mimeType'] = mime_type

        media = MediaFileUpload(file_location)
        file = self.drive_service.files().create(body=file_metadata,
                                                 media_body=media).execute()

        # if this is a service account and auto-ownership is True
        if self.auth_type == 'service_account' and self.auto_owner:
            # move ownership to the default email
            if self.default_owner_email == '':
                raise AttributeError("No default owner email was set.  Cannot change ownership!")
            self.share_document(file_id=file['id'], recipient_email=self.default_owner_email, share_permissions='owner')
        return file

    def get_parent_ids(self, file_id):
        """
        Returns back the parent IDs of the file ID passed in
        :param file_id: Which child file to look up.
        :return: List of parent folder IDs
        """
        if not self.drive_service:
            raise ConnectionError("Drive service has not been turned on or not yet authenticated.  "
                                  "Please run the `authenticate()` function before attempting to manage files.")

        if not file_id:
            raise AttributeError("File ID is required to lookup.")

        parents = self.drive_service.files().get(fileId=file_id,
                                                 fields='parents').execute()

        return parents['parents']

    def append_rows_to_spreadsheet(self, spreadsheet_id, row_data, tab_name=None, tab_index=None, starting_range='A1',
                                   input_type='USER_ENTERED'):
        """
        Takes raw row data and appends it to the bottom of a Google Sheet.
        :param spreadsheet_id: File ID of the spreadsheet work book
        :param row_data: A list of rows such that each row consists of a list of values.  Example: [[1, 2], [3, 4]]
                            would be two rows with 1 and 2 in columns A and B, and a second row with 3 and 4 in columns
                            A and B.
        :param tab_name: If you want to append to a different tab than the first/default, specify the tab name here.
        :param tab_index: If you want to append toa  different tab by index, you can specify that index here.
        :param starting_range: Where to start the insertions.  Default is A1 to specify the append to occur at the last
                                known row, starting in column A.
        :param input_type: Allows you to specify a known INPUT_TYPE_CONSTANT to decide if the values input should be
                            interpreted (as for example, dates, dollars, etc.) or left as "Raw"
        :return: Response from append call
        """
        body = {
            'values': row_data
        }

        range = starting_range

        doc = self.sheet_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = [s['properties'] for s in doc['sheets']]
        sheet = sheets[0]
        if tab_index:
            sheet = sheets[tab_index]
        if tab_name:
            for tab in sheets:
                if tab['title'] == tab_name:
                    sheet = tab

        range = sheet['title']+'!'+range

        # NOTES:
        #   - 'range': Start with the sheet's title to specify which sheet/tab to write these values to, followed by
        #       a '!', and then the starting cell ('A1' for the top-left cell).  Leave as 'A1' for default sheet/tab
        #       (first one).
        #   - 'valueInputOption': 'RAW' forces it to not interpret the text you insert.
        #       See: https://developers.google.com/sheets/api/reference/rest/v4/ValueInputOption
        #   - 'insertDataOption': 'INSERT_ROWS' forces it to append to the bottom of the table.  You can also
        #       'OVERWRITE' which essentially is just a write command. It is safe to leave the `range` set to 'A1'
        #        if this is set to 'INSERT_ROWS' verses 'OVERWRITE'.
        # See: https://developers.google.com/sheets/api/reference/rest/v4/spreadsheets.values/append#InsertDataOption
        response = self.sheet_service.spreadsheets().values().append(spreadsheetId=spreadsheet_id,
                                                                     range=range,
                                                                     valueInputOption=input_type,
                                                                     insertDataOption='INSERT_ROWS',
                                                                     body=body).execute()
        return response

    def append_file_to_spreadsheet(self, spreadsheet_id, csv_file, tab_name=None, tab_index=None, starting_range='A1',
                                   input_type='USER_ENTERED'):
        """
         Takes raw row data and appends it to the bottom of a Google Sheet.
        :param spreadsheet_id: File ID of the spreadsheet work book
        :param csv_file: File containing the row data you want to upload.
        :param tab_name: If you want to append to a different tab than the first/default, specify the tab name here.
        :param tab_index: If you want to append toa  different tab by index, you can specify that index here.
        :param starting_range: Where to start the insertions.  Default is A1 to specify the append to occur at the last
                                known row, starting in column A.
        :param input_type: Allows you to specify a known INPUT_TYPE_CONSTANT to decide if the values input should be
                            interpreted (as for example, dates, dollars, etc.) or left as "Raw"
        :return: Response from append call
        """
        row_data = []
        fin = open(csv_file, 'r')
        csv_reader = csv.reader(fin)
        for row in csv_reader:
            row_data.append(row)

        return self.append_rows_to_spreadsheet(spreadsheet_id=spreadsheet_id, row_data=row_data, tab_name=tab_name,
                                               starting_range=starting_range, input_type=input_type, tab_index=tab_index)

"""
CONSTANTS that can be used throughout various functions
"""
BROWSER_GOOGLE_CHROME = 'chrome'
BROWSER_FIREFOX = 'firefox'

AUTH_TYPE_BROWSER = 'web'
AUTH_TYPE_SERVICE_ACCOUNT = 'service_account'

# Spreadsheet Formats
FORMAT_MS_EXCEL = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
FORMAT_OPEN_OFFICE_SHEET = 'application/x-vnd.oasis.opendocument.spreadsheet'
FORMAT_PDF = 'application/pdf'
FORMAT_CSV = 'text/csv'
FORMAT_TSV = 'text/tab-separated-values'
FORMAT_HTML = 'application/zip'

# Interpretations for data insertion into sheets
INPUT_TYPE_RAW = 'RAW'
INPUT_TYPE_AUTO = 'USER_ENTERED'

# Levels of sharing permissions
PERMISSION_READ = 'reader'
PERMISSION_WRITE = 'writer'
PERMISSION_OWNER = 'owner'
